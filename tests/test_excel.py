"""
测试：Excel 导入导出模块（生成模板、解析数据、API 导入）
"""
import io
import pytest
import openpyxl
from app.services.excel_import import generate_template, parse_excel


class TestGenerateTemplate:
    def test_returns_bytes(self, app, db):
        with app.app_context():
            result = generate_template()
            assert isinstance(result, bytes)
            assert len(result) > 0

    def test_valid_xlsx(self, app, db):
        """生成的文件可以被 openpyxl 正常读取"""
        with app.app_context():
            data = generate_template()
            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            # 第 2 行是表头
            headers = [ws.cell(2, c).value for c in range(1, 9)]
            assert '服务器名称*' in headers
            assert 'IP地址*' in headers

    def test_has_example_row(self, app, db):
        """第 3 行应有示例数据"""
        with app.app_context():
            data = generate_template()
            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            assert ws.cell(3, 1).value is not None  # 示例服务器名称


class TestParseExcel:
    def _make_excel(self, rows):
        """辅助：构造含表头和数据行的 xlsx 字节流"""
        wb = openpyxl.Workbook()
        ws = wb.active
        # 说明行
        ws.append(['说明行，跳过'])
        # 表头行
        ws.append(['服务器名称*', 'IP地址*', '系统类型(linux/windows)*',
                   'SSH端口', '用户名*', '密码', '分组', '备注'])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_parse_valid_data(self, app, db):
        data = self._make_excel([
            ['web01', '192.168.1.1', 'linux', 22, 'root', 'pass', '核心', '备注'],
        ])
        servers, errors = parse_excel(data)
        assert len(errors) == 0
        assert len(servers) == 1
        assert servers[0]['name'] == 'web01'
        assert servers[0]['ip'] == '192.168.1.1'
        assert servers[0]['os_type'] == 'linux'

    def test_parse_windows_type(self, app, db):
        data = self._make_excel([
            ['win-srv', '10.0.0.5', 'windows', 22, 'Administrator', '', '默认分组', ''],
        ])
        servers, errors = parse_excel(data)
        assert len(errors) == 0
        assert servers[0]['os_type'] == 'windows'

    def test_parse_windows_default_admin_user(self, app, db):
        data = self._make_excel([
            ['win-srv2', '10.0.0.6', 'windows', None, None, '', '默认分组', ''],
        ])
        servers, errors = parse_excel(data)
        assert len(errors) == 0
        assert servers[0]['username'] == 'Administrator'

    def test_parse_missing_name(self, app, db):
        data = self._make_excel([
            [None, '192.168.1.2', 'linux', 22, 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert len(errors) > 0
        assert '服务器名称不能为空' in errors[0]

    def test_parse_missing_ip(self, app, db):
        data = self._make_excel([
            ['web02', None, 'linux', 22, 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('IP地址不能为空' in e for e in errors)

    def test_parse_invalid_ip(self, app, db):
        data = self._make_excel([
            ['web02', '999.1.1.1', 'linux', 22, 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('IP地址格式不正确' in e for e in errors)

    def test_parse_missing_username(self, app, db):
        data = self._make_excel([
            ['web03', '192.168.1.3', 'linux', 22, None, '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('用户名不能为空' in e for e in errors)

    def test_parse_invalid_os_type(self, app, db):
        data = self._make_excel([
            ['web04', '192.168.1.4', 'ubuntu', 22, 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('linux、windows 或 macos' in e for e in errors)

    def test_parse_skip_empty_rows(self, app, db):
        data = self._make_excel([
            ['web05', '192.168.1.5', 'linux', 22, 'root', '', '默认分组', ''],
            [None, None, None, None, None, None, None, None],  # 空行
        ])
        servers, errors = parse_excel(data)
        assert len(servers) == 1

    def test_parse_invalid_file(self, app, db):
        _, errors = parse_excel(b'this is not an excel file')
        assert len(errors) > 0
        assert '文件解析失败' in errors[0]

    def test_default_ssh_port(self, app, db):
        """Linux/macOS 连接端口为空时默认为 22"""
        data = self._make_excel([
            ['web06', '192.168.1.6', 'linux', None, 'root', '', '默认分组', ''],
        ])
        servers, errors = parse_excel(data)
        assert len(errors) == 0
        assert servers[0]['ssh_port'] == 22

    def test_parse_invalid_port(self, app, db):
        data = self._make_excel([
            ['web07', '192.168.1.7', 'linux', 70000, 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('连接端口必须在 1-65535 之间' in e for e in errors)

    def test_parse_non_numeric_port(self, app, db):
        data = self._make_excel([
            ['web08', '192.168.1.8', 'linux', 'abc', 'root', '', '默认分组', ''],
        ])
        _, errors = parse_excel(data)
        assert any('第3行' in e and '连接端口格式不正确' in e for e in errors)

    def test_default_windows_winrm_port(self, app, db):
        """Windows 连接端口为空时默认为 5985（WinRM）"""
        data = self._make_excel([
            ['win06', '192.168.1.60', 'windows', None, 'Administrator', '', '默认分组', ''],
        ])
        servers, errors = parse_excel(data)
        assert len(errors) == 0
        assert servers[0]['ssh_port'] == 5985


class TestExcelImportAPI:
    def _make_excel_bytes(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['说明行'])
        ws.append(['服务器名称*', 'IP地址*', '系统类型*', 'SSH端口', '用户名*', '密码', '分组', '备注'])
        ws.append(['api-srv', '172.16.0.1', 'linux', 22, 'root', '', '默认分组', ''])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def test_import_excel_via_api(self, admin_client, db):
        """通过 API 上传 Excel 文件并导入服务器"""
        data = self._make_excel_bytes()
        resp = admin_client.post(
            '/servers/import',
            data={'file': (io.BytesIO(data), 'servers.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert '导入完成'.encode() in resp.data

    def test_import_excel_skip_duplicate_ip(self, admin_client, app, db):
        """重复 IP 应跳过而非报错"""
        # 先导入一次
        data = self._make_excel_bytes()
        admin_client.post(
            '/servers/import',
            data={'file': (io.BytesIO(data), 's.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        # 再导入同一条 → 应跳过
        data2 = self._make_excel_bytes()
        resp = admin_client.post(
            '/servers/import',
            data={'file': (io.BytesIO(data2), 's2.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        assert '跳过重复'.encode() in resp.data

    def test_import_no_file(self, admin_client, db):
        """没有上传文件时应提示错误"""
        resp = admin_client.post(
            '/servers/import',
            data={},
            follow_redirects=True,
        )
        assert '请选择文件'.encode() in resp.data

    def test_import_viewer_forbidden(self, viewer_client, db):
        """viewer 无权导入"""
        resp = viewer_client.post(
            '/servers/import',
            data={},
            follow_redirects=False,
        )
        assert resp.status_code in (301, 302)
