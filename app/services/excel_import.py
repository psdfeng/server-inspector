import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from app.utils.validators import is_valid_ip, is_valid_port


TEMPLATE_HEADERS = ['服务器名称*', 'IP地址*', '系统类型(linux/windows/macos)*', '连接端口(SSH/WinRM)', '用户名*', '密码', '分组', '备注']
TEMPLATE_EXAMPLE = ['Web服务器-01', '192.168.1.100', 'linux', '22', 'root', 'password123', '核心机房', 'Nginx服务器']


def generate_template() -> bytes:
    """生成 Excel 导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '服务器导入模板'
    
    # 标题样式
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    tip_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写入说明行
    ws.merge_cells('A1:H1')
    ws['A1'] = '服务器批量导入模板 - 请勿修改表头，必填项已标注 *，SSH端口默认22'
    ws['A1'].font = Font(bold=True, color='FF0000', size=11)
    ws['A1'].fill = tip_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # 写入表头
    for col, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 30

    # 写入示例数据
    for col, val in enumerate(TEMPLATE_EXAMPLE, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # 调整列宽
    col_widths = [20, 18, 22, 10, 15, 18, 15, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_excel(file_content: bytes) -> tuple:
    """
    解析上传的 Excel 文件
    返回: (servers_list, errors_list)
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
    except Exception as e:
        return [], [f'文件解析失败: {e}']

    servers = []
    errors = []
    
    # 找到数据起始行（跳过说明行和表头）
    for row_no, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue  # 跳过空行
        
        name = str(row[0]).strip() if row[0] else ''
        ip = str(row[1]).strip() if row[1] else ''
        os_type = str(row[2]).strip().lower() if row[2] else 'linux'
        if row[3] is None or str(row[3]).strip() == '':
            ssh_port = 5985 if os_type == 'windows' else 22
            port_parse_error = False
        else:
            try:
                ssh_port = int(str(row[3]).strip())
                port_parse_error = False
            except Exception:
                ssh_port = 0
                port_parse_error = True
        username = str(row[4]).strip() if row[4] else ('Administrator' if os_type == 'windows' else '')
        password = str(row[5]).strip() if row[5] else ''
        group = str(row[6]).strip() if row[6] else '默认分组'
        description = str(row[7]).strip() if row[7] else ''

        row_errors = []
        if not name:
            row_errors.append('服务器名称不能为空')
        if not ip:
            row_errors.append('IP地址不能为空')
        elif not is_valid_ip(ip):
            row_errors.append('IP地址格式不正确')
        if not username:
            row_errors.append('用户名不能为空')
        if port_parse_error:
            row_errors.append('连接端口格式不正确')
        elif not is_valid_port(ssh_port):
            row_errors.append('连接端口必须在 1-65535 之间')
        if os_type not in ('linux', 'windows', 'macos'):
            row_errors.append('系统类型必须为 linux、windows 或 macos')
        
        if row_errors:
            errors.append(f"第{row_no}行: {', '.join(row_errors)}")
        else:
            servers.append({
                'name': name, 'ip': ip, 'os_type': os_type,
                'ssh_port': ssh_port, 'username': username, 'password': password,
                'group': group, 'description': description,
            })

    return servers, errors
